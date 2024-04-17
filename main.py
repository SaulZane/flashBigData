import random
import time
from tkinter import *
from tkinter import messagebox
from openpyxl import load_workbook
import threading
import tkinter.font as tkFont

class ExamSystem:
    def __init__(self, master):
        self.master = master
        self.master.title("考试抽题系统")
        self.master.geometry("1000x500")
        self.font = tkFont.Font(family="黑体", size=14)

        self.questions = self.load_questions_from_excel("题目数据.xlsx")
        self.current_question = None
        self.score = 0
        self.answered_count = 0
        self.start_time = None
        self.elapsed_time = 0

        self.question_label = Label(self.master, wraplength=1000,font=self.font )
        self.question_label.pack(pady=20)

        self.option_buttons = []
        for i in range(4):
            button = Button(self.master, text="", width=1000, command=lambda idx=i: self.submit_answer(chr(ord('A')+idx)),font=self.font)
            button.pack(pady=5)
            self.option_buttons.append(button)

        self.next_button = Button(self.master, text="下一题", width=30, state=DISABLED, command=self.next_question,font=self.font)
        self.next_button.pack(pady=50)

        self.timer_label = Label(self.master, text="倒计时：",font=self.font)
        self.timer_label.pack(pady=10)

        self.score_label = Label(self.master, text="得分：0",font=self.font)
        self.score_label.pack(pady=5)

        self.answered_count_label = Label(self.master, text="已答题数：0/5",font=self.font)
        self.answered_count_label.pack(pady=5)

    def load_questions_from_excel(self, file_path):
        workbook = load_workbook(file_path)
        sheet = workbook.active

        questions = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            question_type = row[0]
            question_text = row[1]
            answer_analysis = row[2]
            correct_answer = row[3]
            options = row[4:]

            question = {
                'type': question_type,
                'text': question_text,
                'analysis': answer_analysis,
                'correct_answer': correct_answer,
                'options': options
            }
            questions.append(question)

        return questions

    def select_question(self):
        self.current_question = random.choice(self.questions)

    def display_question(self):
        self.question_label.config(text=self.current_question['text'])

        if self.current_question['type'] == '单选题':
            options = self.current_question['options']
            for i in range(len(options)):
                self.option_buttons[0].pack(side=TOP, pady=5)
                self.option_buttons[1].pack(side=TOP, pady=5)
                self.option_buttons[2].pack(side=TOP, pady=5)
                self.option_buttons[3].pack(side=TOP, pady=5)
                self.option_buttons[i].config(text=f"{chr(ord('A')+i)}. {options[i]}")
        elif self.current_question['type'] == '判断题':
            self.option_buttons[0].config(text="A. 正确")
            self.option_buttons[1].config(text="B. 错误")
            self.option_buttons[2].config(text="")
            self.option_buttons[3].config(text="")
        for i in range(len(self.option_buttons)):
            self.option_buttons[i].config(state=NORMAL)

    def submit_answer(self, user_answer):
        is_correct = self.check_answer(user_answer)
        self.display_result(is_correct)

        if is_correct:
            self.score += 20

        self.answered_count += 1
        self.score_label.config(text=f"得分：{self.score}")
        self.answered_count_label.config(text=f"已答题数：{self.answered_count}/5")

        self.elapsed_time = time.time() - self.start_time
        if self.elapsed_time > 180 or self.answered_count == 5:
            self.finish_exam()
        else:
            self.next_button.config(state=NORMAL)

    def check_answer(self, user_answer):
        return user_answer == self.current_question['correct_answer']

    def display_result(self, is_correct):
        for i in range(len(self.option_buttons)):
            self.option_buttons[i].config(state=DISABLED)
        if is_correct:
            messagebox.showinfo("回答结果", "回答正确！")
        else:
            messagebox.showinfo("回答结果", "回答错误！\n正确答案是：" + self.current_question['correct_answer'])
            if self.current_question['analysis']:
                messagebox.showinfo("答案解析", self.current_question['analysis'])

    def next_question(self):
        self.next_button.config(state=DISABLED)
        self.select_question()
        self.display_question()

    def start_exam(self):
        self.select_question()
        self.display_question()
        self.start_time = time.time()

        # 创建计时器线程
        timer_thread = threading.Thread(target=self.update_timer)
        timer_thread.daemon = True  # 将线程设置为守护线程，使得主线程退出时自动退出计时器线程
        timer_thread.start()

    def update_timer(self):
        while self.elapsed_time < 180 and self.answered_count < 5:
            remaining_time = 180 - int(self.elapsed_time)
            minutes = remaining_time // 60
            seconds = remaining_time % 60
            self.timer_label.config(text=f"倒计时：{minutes:02d}:{seconds:02d}")
            time.sleep(1)  # 每秒更新一次计时器
            self.elapsed_time = time.time() - self.start_time
            

    def finish_exam(self):
        messagebox.showinfo("考试结束", f"总成绩：{self.score}")
        self.master.quit()

if __name__ == '__main__':
    root = Tk()
    exam_system = ExamSystem(root)
    exam_system.start_exam()
    root.mainloop()